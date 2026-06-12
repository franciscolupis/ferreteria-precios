"""
Exportador de presupuestos a PDF (fpdf2) y Excel (openpyxl).
"""
import io
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)

_CHAR_MAP = str.maketrans({
    "—": "-",   # em dash —
    "–": "-",   # en dash –
    "‒": "-",   # figure dash
    "•": "*",   # bullet •
    "’": "'",   # right single quote '
    "‘": "'",   # left single quote '
    "“": '"',   # left double quote "
    "”": '"',   # right double quote "
    "…": "...", # ellipsis …
    "·": ".",   # middle dot ·
    "→": "->",  # arrow →
    "←": "<-",  # arrow ←
    " ": " ",   # non-breaking space
    "®": "(R)", # registered trademark ®
    "©": "(C)", # copyright ©
    "°": " grados",  # degree °
})


def _sanitizar(texto: str) -> str:
    """Convierte caracteres fuera de Latin-1 a equivalentes ASCII para fpdf Helvetica."""
    texto = str(texto).translate(_CHAR_MAP)
    return texto.encode("latin-1", errors="replace").decode("latin-1")


# ─── PDF ────────────────────────────────────────────────────────────────────

def exportar_pdf(filas: list[dict[str, Any]], titulo: str = "Presupuesto") -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("Instala fpdf2: pip install fpdf2")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _sanitizar(titulo), ln=True, align="C")
    pdf.set_font("Helvetica", size=9)
    pdf.cell(0, 6, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
    pdf.ln(4)

    # Encabezados
    headers = ["Código", "Descripción", "Empaque", "Lista", "Costo+IVA", "Precio Venta", "Ganancia"]
    widths  = [25,       75,            30,        25,      25,          25,              20]
    pdf.set_fill_color(30, 30, 50)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, _sanitizar(h), border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", size=8)
    fill = False
    for f in filas:
        pdf.set_fill_color(240, 240, 248) if fill else pdf.set_fill_color(255, 255, 255)
        datos = [
            str(f.get("codigo_producto") or ""),
            str(f.get("descripcion") or "")[:55],
            str(f.get("empaque") or "-"),
            f"$ {f.get('precio_lista', 0):,.2f}",
            f"$ {f.get('costo_con_iva', 0):,.2f}",
            f"$ {f.get('precio_venta', 0):,.2f}",
            f"$ {f.get('ganancia_neta', 0):,.2f}",
        ]
        for dato, w in zip(datos, widths):
            pdf.cell(w, 6, _sanitizar(dato), border=1, fill=True)
        pdf.ln()
        fill = not fill

    return bytes(pdf.output())


# ─── Excel ──────────────────────────────────────────────────────────────────

def exportar_excel(filas: list[dict[str, Any]], titulo: str = "Presupuesto") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    headers = ["Código", "Descripción", "Empaque", "Precio Lista", "Costo+IVA", "Precio Venta", "Ganancia Neta", "% Ganancia"]
    keys    = ["codigo_producto", "descripcion", "empaque", "precio_lista", "costo_con_iva", "precio_venta", "ganancia_neta", "ganancia_neta_pct"]

    header_fill = PatternFill("solid", fgColor="1E1E32")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    precio_fmt = '#,##0.00'
    pct_fmt    = '0.00"%"'

    for row_idx, f in enumerate(filas, start=2):
        alt_fill = PatternFill("solid", fgColor="F0F0F8") if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(keys, start=1):
            val = f.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if alt_fill:
                cell.fill = alt_fill
            if key in ("precio_lista", "costo_con_iva", "precio_venta", "ganancia_neta"):
                cell.number_format = precio_fmt
                if key == "precio_venta":
                    cell.font = Font(bold=True, color="1B8A4C")
                if key == "ganancia_neta":
                    cell.font = Font(color="C0392B")
            if key == "ganancia_neta_pct":
                cell.number_format = pct_fmt

    col_widths = [12, 50, 18, 15, 15, 15, 15, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
